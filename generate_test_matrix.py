"""
WSW Patch Test Matrix Generator
================================
Creates (or updates) an Excel workbook with two sheets:

  Sheet 1 — "Patch Tests"
    Columns: Patch | Description | <Test1> | <Test2> | ...
    Each row is a patch version. Test cells are left empty for manual tester input.

  Sheet 2 — "Test Definitions"
    Columns: Test | Long Description
    Each row explains what a test column from Sheet 1 means.

If the file already exists, the script:
  - Loads it
  - Adds any NEW test columns to Sheet 1 with a default value of "N/A (added later)"
    for all pre-existing patch rows
  - Adds any NEW test rows to Sheet 2
  - Adds any NEW patch rows to Sheet 1 (empty test cells, ready for the tester)
  - Does NOT overwrite existing cell values

Usage:
    python generate_test_matrix.py
"""

import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════════
# DATA DEFINITIONS — edit these to add patches/tests
# ═══════════════════════════════════════════════════════════════════════

PATCHES = [
    {
        "patch": "0.1.0a",
        "description": (
            "Initial wand suite. 6 wand families × 4 modes (24 items). "
            "Rectangle, Ellipse, Diamond, Triangle, Elbow, CardinalLine shapes. "
            "Filled + Hollow modes. Undo system. Preview overlay. Config panel."
        ),
    },
    {
        "patch": "0.1.1a",
        "description": (
            "Half-Ellipse shapes (H/V). Slice modes. StraightLine shape. "
            "Connect-diameter toggle. Equal-dimensions toggle."
        ),
    },
    {
        "patch": "0.1.2a",
        "description": (
            "Coating Ignore mode (paint, illuminant, echo). Tri-state coating toggles. "
            "8×4 paint grid with Ignore swatch. Vacuum teleport fix. "
            "Progressive batching for wall replacement."
        ),
    },
    {
        "patch": "0.1.3a",
        "description": (
            "MP Day 1-3: Wiring packets, Building packets. "
            "Server-authoritative wiring and building. "
            "Common packet header (22 bytes). OperationResult feedback."
        ),
    },
    {
        "patch": "0.1.4a",
        "description": (
            "MP Day 4: Dismantling packets. Server-authoritative tile/wall destruction. "
            "Server-side pick power validation, container handling, demon altar protection."
        ),
    },
{
    "patch": "0.1.5a",
    "description": (
        "MP Day 5: Replacement packets, Coating packets. "
        "Server-authoritative tile/wall replacement with slope preservation. "
        "Server-authoritative coating (paint/illuminant/echo). "
        "5 coating modes: paint tile, paint wall, scrape paint, scrape moss, harvest moss. "
        "Ignore flags for selective coating preservation."
    ),
},
]

# Test definitions: (short_name, long_description)
# short_name becomes a column header in Sheet 1 and a row label in Sheet 2.
TESTS = [
    (
        "Wand of Building Working",
        (
            "Wand of Building places the selected block in the drawn shape.\n"
            "• The shape draws and applies correctly for all 8 shape types\n"
            "• The wand respects consumption rules and infinite resource config\n"
            "• The wand respects block swap / replace mode\n"
            "• Slope overwrite applies the configured slope correctly\n"
            "• Exhaustion modes work: NextBlock skips, Cancel pre-checks, Interrupt stops\n"
            "• Grass seed placement converts substrates correctly\n"
            "• Wall building path places walls in the shape\n"
            "• Tile wand ammo (e.g., living wood wand consumes wood) works correctly\n"
            "• All 4 selection modes (Instant/Select/Confirm/Stamp) function"
        ),
    ),
    (
        "Wand of Dismantling Working",
        (
            "Wand of Dismantling destroys tiles and/or walls in the drawn shape.\n"
            "• Tiles are destroyed when DestroyTiles is enabled\n"
            "• Walls are destroyed when DestroyWalls is enabled\n"
            "• Both can be enabled simultaneously\n"
            "• Pick power check prevents destroying tiles the player can't mine\n"
            "• CanKillTile validation prevents destroying furniture/support blocks\n"
            "• Top-to-bottom sort handles trees and multi-tile objects correctly\n"
            "• SuppressDrops config prevents item drops when enabled\n"
            "• VacuumItems teleports dropped items to the player when enabled\n"
            "• All 4 selection modes (Instant/Select/Confirm/Stamp) function"
        ),
    ),
    (
        "Wand of Replacement Working",
        (
            "Wand of Replacement swaps source tile types with target types.\n"
            "• Source type detection works for Tile, Platform, Rope, PlanterBox, Wall\n"
            "• Air target acts as selective demolition\n"
            "• Substrate variant detection (grass/jungle grass → dirt/mud) works\n"
            "• WouldTileLoseSupport prevents wall replacement from collapsing torches\n"
            "• Pick power validation for replacement (must be able to mine the source)\n"
            "• Items consumed correctly from inventory\n"
            "• All 4 selection modes function"
        ),
    ),
    (
        "Wand of Wiring Working",
        (
            "Wand of Wiring places or removes wire and actuators in the drawn shape.\n"
            "• All 4 wire colors (red/green/blue/yellow) can be toggled independently\n"
            "• Actuator placement/removal works\n"
            "• Place mode adds wires, Remove mode removes wires\n"
            "• Default shape is Elbow (vanilla wire-kite behavior)\n"
            "• Wire consumption respects infinite resource config\n"
            "• All 4 selection modes function"
        ),
    ),
    (
        "Wand of Safekeeping Working",
        (
            "Wand of Safekeeping protects/unprotects tile positions.\n"
            "• Protect mode marks tiles — other wands refuse to modify them\n"
            "• Unprotect mode clears the protection flag\n"
            "• Protection persists across world saves (TagCompound serialization)\n"
            "• Visual overlay shows protected areas (cyan=tile, magenta=wall, gold=both)\n"
            "• All 4 selection modes function"
        ),
    ),
    (
        "Wand of Coating Working",
        (
            "Wand of Coating applies/removes paint, illuminant, and echo coatings.\n"
            "• All 30 paint colors can be selected and applied\n"
            "• Illuminant toggle applies/removes illuminant coating\n"
            "• Echo toggle applies/removes echo (invisible tile) coating\n"
            "• Target selector: Tiles only, Walls only, Both\n"
            "• Paint consumption respects inventory and infinite resource config\n"
            "• All 4 selection modes function"
        ),
    ),
    (
        "Ignore Coating Mode",
        (
            "Coating wand Ignore mode preserves existing coatings.\n"
            "• Paint Ignore (byte 255): does not overwrite existing paint\n"
            "• Illuminant Ignore: does not change illuminant state\n"
            "• Echo Ignore: does not change echo state\n"
            "• Tri-state UI cycle: Apply (yellow/blue) → Remove (red) → Ignore (gray)\n"
            "• Ignore swatch (dash icon) appears at grid position 31\n"
            "• When all three are Ignore, operation effectively does nothing"
        ),
    ),
    (
        "Shape Drawing Correct",
        (
            "All 8 shape types rasterize correctly.\n"
            "• Rectangle: bounding-box fill, no gaps\n"
            "• Ellipse: smooth Math.Sqrt rasterization, no jagged edges\n"
            "• Diamond: Manhattan distance, symmetric on all 4 quadrants\n"
            "• Triangle: scanline fill handles degenerate cases (1-wide, flat)\n"
            "• Elbow: L-shaped joint, correct axis from first mouse movement\n"
            "• CardinalLine: 8-direction Bresenham-like, no flicker\n"
            "• HalfEllipse H/V: correct half of doubled ellipse\n"
            "• StraightLine: arbitrary angle line\n"
            "• Filled and Hollow modes work for each shape\n"
            "• Equal Dimensions toggle produces perfect circles/squares"
        ),
    ),
    (
        "Preview Overlay Renders",
        (
            "Shape preview overlay renders correctly while selecting.\n"
            "• Overlay appears immediately on mouse drag/click\n"
            "• W×H dimension label displayed correctly\n"
            "• 3 render modes work: AlwaysFullShape, OnlyOutline, OutlineAndPartialFill\n"
            "• Screen culling: no rendering for off-screen tiles\n"
            "• Color coding matches wand family (green=building, red=dismantling, etc.)\n"
            "• Overlay disappears on cancel or execution"
        ),
    ),
    (
        "Undo System Working",
        (
            "Undo system reverts tile operations.\n"
            "• Backspace keybind triggers undo for the last operation\n"
            "• Up to 20 undo steps stored per player\n"
            "• TileSnapshot captures: HasTile, TileType, Frame, Slope, WallType\n"
            "• Batch frame update after undo (FinalizeBatch) updates visual state\n"
            "• Undo works for all wand types (Build, Dismantle, Replace, Coat)\n"
            "• Undo stack is per-player (doesn't affect other players' stacks)"
        ),
    ),
    (
        "Selection Modes Correct",
        (
            "All 4 selection modes work correctly for every wand.\n"
            "• Instant: click-drag, release to execute — no lingering selection\n"
            "• Select: click start → click end → execute immediately\n"
            "• Confirm: click start → click end → click confirm (3 steps)\n"
            "• Stamp: click start → click end → click repeatedly to stamp\n"
            "• Right-click cancels active selection with fade-out animation\n"
            "• Cancel animation shows correct wand-family color\n"
            "• Mode cycling via right-click in inventory works"
        ),
    ),
    (
        "Config Options Respected",
        (
            "Config settings affect wand behavior as documented.\n"
            "• InfiniteResource: blocks/walls/wires not consumed when threshold met\n"
            "• SuppressDrops: destroyed tiles don't drop items\n"
            "• BypassPickaxePower: ignores pick power requirements\n"
            "• AllowDemonAltarDestruction: enables/disables altar breaking\n"
            "• AutoOpenChestsOnDestruction: auto-unlocks chests before destroying\n"
            "• VacuumItems: teleports dropped items to player\n"
            "• SelectionCap (Big/Small/Hollow): limits shape size\n"
            "• ProgressiveMode: enables batched execution\n"
            "• MaxOutlineThickness: limits hollow thickness slider"
        ),
    ),
    (
        "Container Destruction",
        (
            "Wand of Dismantling handles containers correctly.\n"
            "• Chests with items: contents drop as item entities before tile destruction\n"
            "• Locked chests: only destroyed if AutoOpenChestsOnDestruction is ON\n"
            "• Dressers: treated as containers, contents dropped\n"
            "• Barrels: treated as containers\n"
            "• Container pass happens BEFORE tile pass (Chest.FindChest needs tiles)\n"
            "• Undo snapshots cover container tiles\n"
            "• SuppressDrops also affects container content drops"
        ),
    ),
    (
        "Demon Altar Protection",
        (
            "Demon Altars are protected by default.\n"
            "• AllowDemonAltarDestruction = false: altars are never destroyed\n"
            "• AllowDemonAltarDestruction = true: requires hammer power ≥ 80\n"
            "• BypassPickaxePower + AllowDemonAltarDestruction: can destroy with any tool\n"
            "• Protection applies in both dismantling and replacement wands"
        ),
    ),
    (
        "Progressive Mode Working",
        (
            "Large operations are split into timed batches.\n"
            "• Operations larger than batch size are split into waves\n"
            "• Each wave processes ProgressiveBatchSize tiles\n"
            "• ProgressiveInterval seconds between each wave\n"
            "• Status message shows tile count, wave count, and estimated time\n"
            "• Tiles are destroyed with natural sound/dust/gore effects per wave\n"
            "• Undo captures full operation across all waves\n"
            "• New operation cancels pending progressive waves"
        ),
    ),
    (
        "Safekeeping Persists",
        (
            "Protection data survives world save/load.\n"
            "• Protected tiles remain protected after saving and reloading the world\n"
            "• TagCompound serialization includes all protected positions\n"
            "• Visual overlay reappears after reload\n"
            "• Other wands still respect protection after reload"
        ),
    ),
    (
        "Thickness & Hollow Mode",
        (
            "Hollow mode and thickness controls work correctly.\n"
            "• Thickness 0: slim 1-pixel boundary (4-connected)\n"
            "• Thickness 1: standard boundary (8-connected)\n"
            "• Thickness 2+: progressive Chebyshev erosion creates thicker outlines\n"
            "• '[' and ']' keybinds adjust thickness in real-time\n"
            "• MaxOutlineThickness config limits the slider\n"
            "• Hollow mode works for all 8 shape types"
        ),
    ),
    (
        "MP Wiring Sync",
        (
            "Wiring operations synchronize in multiplayer.\n"
            "• Player A places wires → Player B sees them appear\n"
            "• Player A removes wires → Player B sees them disappear\n"
            "• All 4 wire colors + actuators sync correctly\n"
            "• Safekeeping protection blocks wiring on server side\n"
            "• Distance cap enforced server-side"
        ),
    ),
    (
        "MP Building Sync",
        (
            "Building operations synchronize in multiplayer.\n"
            "• Player A builds tiles → Player B sees them appear\n"
            "• Player A builds walls → Player B sees them appear\n"
            "• Inventory consumption syncs (SyncEquipment)\n"
            "• Exhaustion modes work server-side (Cancel/Interrupt)\n"
            "• Replace mode validated server-side (pick power check)\n"
            "• Slope overwrite syncs correctly\n"
            "• Distance cap enforced"
        ),
    ),
    (
        "MP Dismantling Sync",
        (
            "Dismantling operations synchronize in multiplayer.\n"
            "• Player A destroys tiles → Player B sees them disappear\n"
            "• Player A destroys walls → Player B sees them disappear\n"
            "• SuppressDrops config respected server-side\n"
            "• Pick power validated server-side\n"
            "• Demon altar protection enforced server-side\n"
            "• Container destruction syncs correctly\n"
            "• Distance cap enforced"
        ),
    ),
    (
        "Keybinds Functional",
        (
            "All 4 keybinds work correctly.\n"
            "• '[' decreases hollow thickness\n"
            "• ']' increases hollow thickness\n"
            "• '.' toggles the settings panel for the held wand\n"
            "• 'Backspace' triggers undo step\n"
            "• Keybinds are rebindable in settings\n"
            "• Keybinds only activate when a wand is held"
        ),
    ),
    (
        "Settings UI Panels",
        (
            "Per-wand settings panels display and function correctly.\n"
            "• Panel opens on right-click (not selecting) or '.' keybind\n"
            "• Panel is draggable\n"
            "• Shape selector shows all available shapes with icons\n"
            "• Fill mode toggle (Filled/Hollow) works\n"
            "• Slice mode dropdown works\n"
            "• Object type / place type selectors work per wand\n"
            "• Settings persist while holding the wand\n"
            "• Panel closes when switching to a non-wand item"
        ),
    ),
    (
        "Lore Tooltips",
        (
            "Wand tooltips display correctly.\n"
            "• Shift-gated lore text appears only when holding Shift\n"
            "• Shared + wand-specific lore lines display\n"
            "• Config toggle to disable lore works\n"
            "• Tooltip shows wand mode (Instant/Select/Confirm/Stamp)\n"
            "• Tooltip shows current shape and settings summary"
        ),
    ),
    (
        "Recipes & Crafting",
        (
            "All wand recipes are correct and functional.\n"
            "• Each wand family has one craftable item (Instant mode)\n"
            "• Other modes obtained via right-click cycling in inventory\n"
            "• Recipes use appropriate ingredients (gems, bars, etc.)\n"
            "• Crafting stations are correct (Anvils, workbench)\n"
            "• Shimmer results defined for cycling variants"
        ),
    ),
    (
        "No SP Regression",
        (
            "Single-player functionality has no regressions after MP changes.\n"
            "• All wands work identically to pre-MP behavior\n"
            "• No stray network messages in SP (netMode == 0)\n"
            "• Performance is not degraded\n"
            "• gen-skip hack removed without side effects\n"
            "• Undo still works in SP"
        ),
    ),
(
    "MP Replacement Sync",
    (
        "Replacement operations synchronize in multiplayer.\n"
        "• Player A replaces tiles → Player B sees the swap\n"
        "• Player A replaces walls → Player B sees the swap\n"
        "• Source type detection (tile variants like grass→dirt) works server-side\n"
        "• Slope and half-block states preserved after replacement\n"
        "• Pick power validated server-side (must be able to mine source)\n"
        "• WouldTileLoseSupport check prevents collapsing torches on wall replace\n"
        "• Item consumption syncs correctly\n"
        "• Air target (selective demolition) works in MP\n"
        "• Distance cap enforced server-side"
    ),
),
(
    "MP Coating Sync",
    (
        "Coating operations synchronize in multiplayer.\n"
        "• Player A paints tiles → Player B sees the color\n"
        "• Player A paints walls → Player B sees the color\n"
        "• Illuminant coating syncs (tiles glow for both players)\n"
        "• Echo coating syncs (tiles become invisible for both players)\n"
        "• Scrape paint mode removes coatings for all players\n"
        "• Scrape moss converts moss tiles server-side\n"
        "• Harvest moss (LongMoss) works server-side\n"
        "• Ignore flags (paint/illuminant/echo) respected server-side\n"
        "• Paint consumption syncs correctly\n"
        "• Distance cap enforced server-side"
    ),
),
]

# ═══════════════════════════════════════════════════════════════════════
# STYLE DEFINITIONS
# ═══════════════════════════════════════════════════════════════════════

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

PATCH_FONT = Font(name="Calibri", bold=True, size=11)
PATCH_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

TEST_DEF_HEADER_FILL = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
TEST_DEF_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

WRAP_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
NA_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
NA_FONT = Font(name="Calibri", italic=True, color="999999", size=10)

OUTPUT_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "WSW_Test_Matrix.xlsx",
)

RETROACTIVE_DEFAULT = "N/A (added later)"


def build_workbook():
    """Create the workbook from scratch."""
    wb = Workbook()

    # ── Sheet 1: Patch Tests ──────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Patch Tests"
    ws1.sheet_properties.tabColor = "2F5496"

    test_names = [t[0] for t in TESTS]
    headers = ["Patch", "Description"] + test_names

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws1.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Write patch rows
    for row_idx, patch in enumerate(PATCHES, start=2):
        ws1.cell(row=row_idx, column=1, value=patch["patch"]).font = PATCH_FONT
        ws1.cell(row=row_idx, column=1).fill = PATCH_FILL
        ws1.cell(row=row_idx, column=1).border = THIN_BORDER
        ws1.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="top")

        desc_cell = ws1.cell(row=row_idx, column=2, value=patch["description"])
        desc_cell.alignment = WRAP_ALIGNMENT
        desc_cell.border = THIN_BORDER

        # Test columns — leave empty for tester to fill
        for col_idx in range(3, len(headers) + 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value="")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

    # Column widths
    ws1.column_dimensions["A"].width = 12
    ws1.column_dimensions["B"].width = 55
    for col_idx in range(3, len(headers) + 1):
        ws1.column_dimensions[get_column_letter(col_idx)].width = 16

    # Freeze panes: lock headers + Patch/Description columns
    ws1.freeze_panes = "C2"

    # ── Sheet 2: Test Definitions ─────────────────────────────────
    ws2 = wb.create_sheet(title="Test Definitions")
    ws2.sheet_properties.tabColor = "548235"

    # Headers
    for col_idx, header in enumerate(["Test", "Long Description"], start=1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = TEST_DEF_HEADER_FONT
        cell.fill = TEST_DEF_HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Test rows
    for row_idx, (name, desc) in enumerate(TESTS, start=2):
        name_cell = ws2.cell(row=row_idx, column=1, value=name)
        name_cell.font = Font(name="Calibri", bold=True, size=11)
        name_cell.alignment = Alignment(vertical="top")
        name_cell.border = THIN_BORDER

        desc_cell = ws2.cell(row=row_idx, column=2, value=desc)
        desc_cell.alignment = WRAP_ALIGNMENT
        desc_cell.border = THIN_BORDER

    # Column widths
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 80

    ws2.freeze_panes = "A2"

    return wb


def update_workbook(wb):
    """Update an existing workbook: add missing tests/patches without overwriting data."""
    ws1 = wb["Patch Tests"]
    ws2 = wb["Test Definitions"]

    test_names = [t[0] for t in TESTS]

    # ── Ensure base headers exist on Sheet 1 ──────────────────────
    if ws1.cell(row=1, column=1).value != "Patch":
        cell = ws1.cell(row=1, column=1, value="Patch")
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    if ws1.cell(row=1, column=2).value != "Description":
        cell = ws1.cell(row=1, column=2, value="Description")
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    ws1.column_dimensions["A"].width = 12
    ws1.column_dimensions["B"].width = 55

    # ── Ensure base headers exist on Sheet 2 ──────────────────────
    if ws2.cell(row=1, column=1).value != "Test":
        cell = ws2.cell(row=1, column=1, value="Test")
        cell.font = TEST_DEF_HEADER_FONT
        cell.fill = TEST_DEF_HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    if ws2.cell(row=1, column=2).value != "Long Description":
        cell = ws2.cell(row=1, column=2, value="Long Description")
        cell.font = TEST_DEF_HEADER_FONT
        cell.fill = TEST_DEF_HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 80

    # ── Discover existing state ───────────────────────────────────
    existing_headers = []
    for col_idx in range(1, ws1.max_column + 1):
        val = ws1.cell(row=1, column=col_idx).value
        if val:
            existing_headers.append(val)

    existing_patches = set()
    for row_idx in range(2, ws1.max_row + 1):
        val = ws1.cell(row=row_idx, column=1).value
        if val:
            existing_patches.add(str(val).strip())

    existing_test_defs = set()
    for row_idx in range(2, ws2.max_row + 1):
        val = ws2.cell(row=row_idx, column=1).value
        if val:
            existing_test_defs.add(str(val).strip())

    # ── Add missing test COLUMNS to Sheet 1 ───────────────────────
    for test_name in test_names:
        if test_name not in existing_headers:
            new_col = ws1.max_column + 1

            cell = ws1.cell(row=1, column=new_col, value=test_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

            for row_idx in range(2, ws1.max_row + 1):
                patch_val = ws1.cell(row=row_idx, column=1).value
                if patch_val:
                    na_cell = ws1.cell(row=row_idx, column=new_col, value=RETROACTIVE_DEFAULT)
                    na_cell.fill = NA_FILL
                    na_cell.font = NA_FONT
                    na_cell.alignment = Alignment(horizontal="center", vertical="center")
                    na_cell.border = THIN_BORDER

            ws1.column_dimensions[get_column_letter(new_col)].width = 16
            existing_headers.append(test_name)

    # ── Add missing PATCH ROWS to Sheet 1 ─────────────────────────
    for patch in PATCHES:
        if patch["patch"] not in existing_patches:
            new_row = ws1.max_row + 1

            ws1.cell(row=new_row, column=1, value=patch["patch"]).font = PATCH_FONT
            ws1.cell(row=new_row, column=1).fill = PATCH_FILL
            ws1.cell(row=new_row, column=1).border = THIN_BORDER
            ws1.cell(row=new_row, column=1).alignment = Alignment(horizontal="center", vertical="top")

            desc_cell = ws1.cell(row=new_row, column=2, value=patch["description"])
            desc_cell.alignment = WRAP_ALIGNMENT
            desc_cell.border = THIN_BORDER

            for col_idx in range(3, len(existing_headers) + 1):
                cell = ws1.cell(row=new_row, column=col_idx, value="")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = THIN_BORDER

    # ── Add missing test ROWS to Sheet 2 ──────────────────────────
    for test_name, test_desc in TESTS:
        if test_name not in existing_test_defs:
            new_row = ws2.max_row + 1

            name_cell = ws2.cell(row=new_row, column=1, value=test_name)
            name_cell.font = Font(name="Calibri", bold=True, size=11)
            name_cell.alignment = Alignment(vertical="top")
            name_cell.border = THIN_BORDER

            desc_cell = ws2.cell(row=new_row, column=2, value=test_desc)
            desc_cell.alignment = WRAP_ALIGNMENT
            desc_cell.border = THIN_BORDER

    # ── Set freeze panes ──────────────────────────────────────────
    ws1.freeze_panes = "C2"
    ws2.freeze_panes = "A2"
def main():
    if os.path.exists(OUTPUT_PATH):
        print(f"Updating existing workbook: {OUTPUT_PATH}")
        wb = load_workbook(OUTPUT_PATH)

        # Ensure both sheets exist
        if "Patch Tests" not in wb.sheetnames:
            wb.create_sheet("Patch Tests", 0)
        if "Test Definitions" not in wb.sheetnames:
            wb.create_sheet("Test Definitions")

        update_workbook(wb)
    else:
        print(f"Creating new workbook: {OUTPUT_PATH}")
        wb = build_workbook()

    wb.save(OUTPUT_PATH)
    print(f"Saved: {OUTPUT_PATH}")
    print(f"  Patches: {len(PATCHES)}")
    print(f"  Tests:   {len(TESTS)}")


if __name__ == "__main__":
    main()
